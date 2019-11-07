import openpyxl
import datetime

workbook = openpyxl.load_workbook("input_FileName.xlsx")
sheet1 = workbook["Sheet1"]
sheet2 = workbook["Sheet2"]

dates = []
days = []
members = []

for i in range(3, sheet1.max_column):
    cell_value = sheet1.cell(row = 1, column = i).value
    if cell_value not in dates:
        dates.append(cell_value)
        print("cell_value:" + str(cell_value))
    
    cell_value = sheet1.cell(row = 2, column = i).value
    if cell_value != '':
        days.append(cell_value)
for i in range(3, sheet2.max_row+1):
    cell_value = sheet2.cell(row = i, column = 1).value
    if cell_value != '':
        members.append(cell_value)


print(dates)
print(days)
cell_column = 2
for date in dates:
    #sheet2.cell(row = 1,  column = cell_column).value = date
    cell_column += 1
cell_column = 2
for day in days:
    sheet2.cell(row = 2, column = cell_column).value = day
    cell_column += 1
print(members)

#Pick up shift day & member
shift_name = ''
shift_names = []
shift_day = []
shift_print = []
count = 2
cell_row = 3
print("sheet1.max_column:" + str(sheet1.max_column))
print("sheet1.max_row:" + str(sheet1.max_row))
for name in members:
    #for j in range(3, 10):
    for j in range(3, sheet1.max_column+1):
        for i in range(3, sheet1.max_row):
            count += 1
            shift_name =  sheet1.cell(row = i, column = j).value
            shift_names.append(shift_name)
            if shift_name == '':
                break

            if shift_name == name:
                shift_day.append("〇")
                break
            elif count == sheet1.max_column - 2:
                shift_day.append("")
        
        count = 0
    print(shift_day)   
    cell_column = 2
    for write in shift_day:
        sheet2.cell(row = cell_row,  column = cell_column).value = write
        cell_column += 1
    cell_row += 1
    shift_day = []     

workbook.save("output_FileName.xlsx")
