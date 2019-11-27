import openpyxl
#Excelファイルを開く
workbook = openpyxl.load_workbook("./test.xlsx")
sheet = workbook["Sheet1"]
#for(×;×;〇)
print(sheet.max_column)
print(sheet.max_row)
#for 変数名　range(始値,終値)
for i in range(1,sheet.max_column+1):
    for j in range(1, sheet.max_row+1):
        #cell_value変数にrow = j,column = iを指定して代入する
        cell_value = sheet.cell(row = j, column = i).value
        print("cell_value:" + str(cell_value))

#use list
#C言語における配列はPythonにおいてはリストとして表現する

#リストの初期化
cell_list = []
print(cell_list) #[]
#リストに追加
cell_list.append(5)
print(cell_list) #[5]

cell_list = []
print(cell_list) #[]

cell_line = []
#上部と組み合わせる
#for 変数名　range(始値,終値)
for i in range(1,sheet.max_column+1):
    for j in range(1, sheet.max_row+1):
        #cell_value変数にrow = j,column = iを指定して代入する
        cell_value = sheet.cell(row = j, column = i).value
        cell_line.append(cell_value)#列の抽出
        #print("cell_value:" + str(cell_value))
    cell_list.append(cell_line)#列を加える
    cell_line = []#次の列を読み取るために初期化

print(cell_list)#全リストの表示

#リストの個別表示
#inの後にリストが来るとき、リストの値をin前の変数に入れる処理をしてくれる
#最初のfor文で二次元リスト([][])の手前側を指定して、次のfor文で後ろを指定する
for line in cell_list: 
    for value in line:
        print(value)

#Excelへの値の追加
#空きcellと行の判定
brank_row = 0
for i in range(sheet.max_row,30):
    brank = sheet.cell(row = i, column = 1).value
    print(brank)
    if brank == None:
        break
brank_row = i
print(brank_row)

#空きcellと列の判定
brank_column = 0
for i in range(sheet.max_column,30):
    brank = sheet.cell(row = 1, column = i).value
    print(brank)
    if brank == None:
        break
    brank = ""
brank_column = i
print(brank_column)
#celへの記述 失敗例 実行してみれば分かる
for line in cell_list:
    for value in line:
        sheet.cell(row = brank_row, column = brank_column).value = value
        brank_row += 1
    brank_column += 1
#原因：最終行と最終列を指定したため、ずれにずれまくった
#iで行番号を指定して、随時増やすことで解決
#成功例
#for line in cell_list:
#    i = 1
#    for value in line:
#        sheet.cell(row = i, column = brank_column).value = value
#        i += 1
#    brank_column += 1

#エクセルファイルの保存 Excelファイルが開かれているとエラー
workbook.save('./test.xlsx')

#PythonにおけるExcel操作の基本はこれでOKなはず
#取得と記述、簡単なリストの使い方は出来たと思う
#ファイルのアクセス方法とかは福岡さんの方が詳しいと思うからとっ捕まえて聞くがよろし